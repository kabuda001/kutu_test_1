import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QLineEdit, QPushButton, QFileDialog, QMessageBox,QTextEdit
from openpyxl import load_workbook


class ExcelParser(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle('订单编号提取')
        self.setGeometry(100, 100, 400, 150)

        # 创建布局
        layout = QVBoxLayout()

        # 创建 QLineEdit 用于显示或输入 Excel 文件路径
        self.file_input = QLineEdit(self)
        self.file_input.setPlaceholderText("请输入或选择Excel文件路径")

        # 创建选择文件按钮
        self.select_button = QPushButton("选择文件", self)
        self.select_button.clicked.connect(self.select_file)  # 按钮点击时调用选择文件方法

        # 创建确认按钮
        self.confirm_button = QPushButton("确认", self)
        self.confirm_button.clicked.connect(self.parse_excel)  # 确认按钮点击时解析 Excel

        # 创建取消按钮
        self.cancel_button = QPushButton("取消", self)
        self.cancel_button.clicked.connect(self.close)  # 关闭窗口

        # 用于显示提取的订单号列表
        self.result_display = QTextEdit(self)
        self.result_display.setReadOnly(True)  # 设置为只读，用户可以复制内容

        # 将控件添加到布局
        layout.addWidget(self.file_input)
        layout.addWidget(self.select_button)
        layout.addWidget(self.confirm_button)
        layout.addWidget(self.cancel_button)
        layout.addWidget(self.result_display)

        # 设置布局
        self.setLayout(layout)

    def select_file(self):
        # 弹出文件选择对话框，选择 Excel 文件
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", "Excel Files (*.xls *.xlsx)")

        # 如果用户选择了文件，将路径显示到 QLineEdit 中
        if file_path:
            self.file_input.setText(file_path)

    def parse_excel(self):
        # 获取文件路径
        file_path = self.file_input.text()

        # 检查路径是否为空
        if not file_path:
            self.show_message("错误", "请提供一个有效的文件路径")
            return

        try:
            # 使用 openpyxl 读取 Excel 文件
            workbook = load_workbook(file_path)
            sheet = workbook.active  # 获取当前活动工作表

            # 查找"订单号"列索引
            order_numbers = []
            header = [cell.value for cell in sheet[1]]  # 获取第一行作为列标题

            if "订单号" in header:
                order_index = header.index("订单号") + 1  # 获取"订单号"列的索引（openpyxl是1-based索引）

                # 遍历"订单号"列的所有数据，提取非空的值
                for row in sheet.iter_rows(min_row=2, min_col=order_index, max_col=order_index):  # 从第二行开始遍历
                    if row[0].value:
                        order_numbers.append(str(row[0].value))

                # 将订单号数据用逗号分隔
                result_str = ', '.join(order_numbers)
                self.result_display.setText(result_str)
            elif "订单编号" in header:
                order_index = header.index("订单编号") + 1  # 获取"订单号"列的索引（openpyxl是1-based索引）

                # 遍历"订单号"列的所有数据，提取非空的值
                for row in sheet.iter_rows(min_row=2, min_col=order_index, max_col=order_index):  # 从第二行开始遍历
                    if row[0].value:
                        order_numbers.append(str(row[0].value))

                # 将订单号数据用逗号分隔
                result_str = ', '.join(order_numbers)
                self.result_display.setText(result_str)
            else:
                self.show_message("错误", "Excel文件中没有找到'订单号'或者'订单编号'这一列")

        except Exception as e:
            # 如果解析 Excel 失败，弹出错误信息
            self.show_message("错误", f"解析文件失败: {str(e)}")

    def show_message(self, title, message):
        # 弹出消息框显示信息
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec_()


if __name__ == '__main__':
    app = QApplication(sys.argv)

    window = ExcelParser()
    window.show()

    sys.exit(app.exec_())
