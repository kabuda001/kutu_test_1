import os
import sys


from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QHBoxLayout, QLabel,QMessageBox
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl,re,shutil

class LoadThread(QThread):
    # 定义信号，当任务完成时发送消息
    finished_signal = pyqtSignal(str)

    def __init__(self, folder1,order_file):
        super().__init__()
        self.folder1 = folder1
        self.order_file = order_file
        self.lack_package = os.path.join(os.path.dirname(order_file), '缺cdr图.xlsx')
        self.error_package = os.path.join(os.path.dirname(order_file), '解析异常的订单.xlsx')

    def run(self):
        # 在这个方法中执行耗时的任务
        try:
            print(f"开始加载文件夹1: {self.folder1}")
            print(f"开始加载订单列表: {self.order_file}")



            cdr_files_map = self.get_cdr_files_map(self.folder1)
            # # 打印字典内容
            # for file_name, file_path in cdr_files_map.items():
            #     print(f"文件名: {file_name}, 文件路径: {file_path}")
            self.read_excel(order_file=self.order_file)
            # import time
            # time.sleep(5)  # 模拟一个耗时任务
            # key 为excel第一行的值
            excel_data = self.read_excel(order_file=self.order_file)
            for row_data in excel_data.values():
                try:
                    self.handleRow(cdr_files_map,row_data)
                except Exception as e:
                    self.appendRow(row_data,self.error_package)
            # 完成后，发送信号到主线程
            self.finished_signal.emit("处理完毕!")
        except Exception as e:
            self.finished_signal.emit(f"处理失败: {str(e)}")

    def is_empty_string(self,value):
        # 检查是否是字符串类型且为空字符串
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    def is_valid_longest_side(self,longest_side):
        # 定义允许的数值集合
        valid_values = {15, 20, 30, 40, 50, 60, 70, 80, 90, 100}

        # 判断 longest_side 是否是数字
        if isinstance(longest_side, (int, float)):  # 如果是整数或浮点数
            return longest_side in valid_values
        elif isinstance(longest_side, str) and longest_side.isdigit():  # 如果是字符串且为数字
            return int(longest_side) in valid_values
        else:
            return False

    # 处理每一行
    def handleRow(self, cdr_files_map, row_data):
        order_num = row_data.get('订单编号')
        if not order_num:
            order_num = row_data.get('订单号')
        if not order_num:
            self.appendRow(row_data, self.error_package)
            return
        specification_name_str = row_data.get('规格名称')
        style, longest_side = self.parse_specification_name_str(specification_name_str)
        cdr_file_path = cdr_files_map.get(style)
        if not cdr_file_path:
            # 款号不存在
            self.appendCdrRow(row_data,style, longest_side, self.lack_package)
            return
        if not self.is_valid_longest_side(longest_side):
            # 最长边解析异常
            self.appendRow(row_data, self.error_package)
            return

    def parse_specification_name_str(self,specification_name_str):
        # 提取款号 (假设款号是字母+数字的组合，可以匹配如 CD115-A)
        match_model = re.search(r'([A-Za-z]+\d+-[A-Za-z])', specification_name_str)
        # 提取所有尺寸，可能的格式为：40cm高x35cm宽 或 60x60cm
        size_matches = re.findall(r'(\d+)(?=cm|x|CM|厘米|\*|X|公分)', specification_name_str)
        if match_model and size_matches:
            model_number = match_model.group(1)
            # 转换尺寸为整数并找到最大的尺寸作为最长边
            sizes = [int(size) for size in size_matches]
            max_size = max(sizes)
        return model_number,max_size

    def appendRow(self, row_data,file_path):
        if os.path.exists(file_path):
            # 文件已存在，打开已有文件
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            # 文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            sheet = wb.active
            # 写入标题行（假设 data 的第一个字典是列名）
            headers = list(row_data.keys())
            sheet.append(headers)
        values = list(row_data.values())
        sheet.append(values)
        # 保存文件
        wb.save(file_path)

    def appendCdrRow(self, row_data,style, longest_side, file_path):
        num = row_data.get('数量')
        if not num:
            num = row_data.get('商品数量')
        order_num = row_data.get('订单编号')
        if not order_num:
            order_num = row_data.get('订单号')
        new_row_data  = {
            '订单编号': order_num,
            '店铺名称': row_data.get('店铺名称'),
            '规格名称': row_data.get('规格名称'),
            '规格':style,
            '最长边':longest_side,
            '数量':num,
            '总价': row_data.get('总价'),
            '实收': row_data.get('实收')

        }
        if os.path.exists(file_path):
            # 文件已存在，打开已有文件
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            # 文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            sheet = wb.active
            # 写入标题行（假设 data 的第一个字典是列名）
            headers = list(new_row_data.keys())
            sheet.append(headers)
        values = list(new_row_data.values())
        sheet.append(values)
        # 保存文件
        wb.save(file_path)

    # 读取订单excel
    def read_excel(self,order_file):
        workbook = openpyxl.load_workbook(order_file)
        # 选择活动工作表
        sheet = workbook.active
        # 读取第一行作为键（keys）
        header = [cell.value for cell in sheet[1]]  # 第一行作为键
        # 读取数据（从第二行开始）
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {header[i]: row[i] for i in range(len(header))}
            data[row_data[header[0]]] = row_data  # 使用第一列的值作为外层字典的 key
        return data

    def is_cdr_file(self,file_path):
        """检查文件是否为 .cdr 格式"""
        return file_path.lower().endswith('.cdr')

    def get_cdr_files_map(self,directory):
        """遍历目录，返回文件名为键，文件路径为值的字典"""
        cdr_files_map = {}

        # 遍历指定目录下的所有文件和文件夹
        for root, dirs, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)

                # 如果是 .cdr 文件，则将文件名和路径存入字典
                if self.is_cdr_file(file_path):
                    file_name =os.path.splitext(os.path.basename(file))[0]  # 获取文件名（不含路径）
                    cdr_files_map[file_name] = file_path

        return cdr_files_map
class FolderSelector(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        # 创建控件
        self.label1 = QLabel('请选择图库文件夹:')

        self.folder1_path = QLabel('未选择文件夹')
        # self.folder1_path = QLabel('//xinguo/贴纸.生产资料/美工做的新款，待检查')


        self.select_folder1_btn = QPushButton('图库文件夹')
        self.ok_btn = QPushButton('确定')
        self.cancel_btn = QPushButton('取消')

        # 按钮点击事件
        self.select_folder1_btn.clicked.connect(self.select_folder1)
        self.ok_btn.clicked.connect(self.ok_clicked)
        self.cancel_btn.clicked.connect(self.cancel_clicked)

        # 布局设置
        layout = QVBoxLayout()

        # 第一个文件夹选择行
        folder1_layout = QHBoxLayout()
        folder1_layout.addWidget(self.label1)
        folder1_layout.addWidget(self.folder1_path)
        folder1_layout.addWidget(self.select_folder1_btn)
        layout.addLayout(folder1_layout)


        # 创建按钮：选择文件
        self.open_button = QPushButton('订单列表，请选择 .xlsx 文件')
        self.open_button.clicked.connect(self.open_file_dialog)
        # 创建标签，用于显示选择的文件路径
        self.order_file = QLabel('未选择文件', self)
        # self.order_file = QLabel('E:/627个.xlsx', self)
        # 将按钮和标签加入布局
        layout.addWidget(self.open_button)
        layout.addWidget(self.order_file)

        # 确定和取消按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.ok_btn)
        buttons_layout.addWidget(self.cancel_btn)
        layout.addLayout(buttons_layout)

        self.setLayout(layout)

        # 窗口设置
        self.setWindowTitle('cdr缺失-zc')
        self.setGeometry(200, 200, 600, 150)

    def open_file_dialog(self):
        # 打开文件选择框，限制只选择 .xlsx 文件
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "选择 .xlsx 文件", "", "Excel 文件 (*.xlsx);;所有文件 (*)",
                                                   options=options)

        if file_name:
            # 显示选择的文件路径
            self.order_file.setText(f"{file_name}")
        else:
            self.order_file.setText("未选择文件")

    def select_folder1(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择文件夹 1')
        if folder_path:
            self.folder1_path.setText(folder_path)


    def ok_clicked(self):
        # 获取选择的文件夹路径
        folder1 = self.folder1_path.text()
        order_file = self.order_file.text()
        if not folder1  or not os.path.exists(folder1):
            QMessageBox.warning(self, '警告', '请确保图库文件夹已经选择且正确！')
            return
        if not order_file or not os.path.isfile(order_file):
            QMessageBox.warning(self, '警告', '订单excel选择有误！')
            return
        # 启动后台线程执行任务
        self.thread = LoadThread(folder1,order_file)
        self.thread.finished_signal.connect(self.on_load_finished)
        self.thread.start()

        # 禁用按钮以防止重复点击
        self.ok_btn.setDisabled(True)
        self.cancel_btn.setDisabled(True)

    def cancel_clicked(self):
        # 创建确认对话框
        reply = QMessageBox.question(self, '确认取消', '你确定要取消操作吗?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        print("操作已取消")
        self.close()

    def on_load_finished(self, message):
        # 加载完成后弹出提示框
        QMessageBox.information(self, '提示', message)

        # 启用按钮
        self.ok_btn.setEnabled(True)
        self.cancel_btn.setEnabled(True)

        # 如果操作完成，关闭窗口
        self.close()


def main():
    app = QApplication(sys.argv)
    window = FolderSelector()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
