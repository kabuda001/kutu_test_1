import os
import sys

from Demos.OpenEncryptedFileRaw import dst_dir
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QHBoxLayout, QLabel,QMessageBox, QRadioButton, QGroupBox
from PyQt5.QtCore import QThread, pyqtSignal
import openpyxl,re,shutil
import win32com.client



class LoadThread(QThread):
    # 定义信号，当任务完成时发送消息
    finished_signal = pyqtSignal(str)

    def __init__(self, folder1, folder2,order_file,copy_true):
        super().__init__()
        self.folder1 = folder1
        self.folder2 = folder2
        self.order_file = order_file
        self.copy_true = copy_true
        self.lack_package = os.path.join(self.folder2, '缺cdr图.xlsx')
        self.error_package = os.path.join(self.folder2, '解析异常的订单.xlsx')
        self.multiple_order_package = os.path.join(self.folder2, '包含多件的订单.xlsx')
        self.remain_package = os.path.join(self.folder2, '有备注的订单.xlsx')

    def run(self):
        # 在这个方法中执行耗时的任务
        try:
            print(f"开始加载文件夹1: {self.folder1}")
            print(f"开始加载文件夹2: {self.folder2}")
            print(f"开始加载订单列表: {self.order_file}")


            # 假设耗时任务是在这儿
            # 这里可以放置您的加载操作代码，比如读取文件，处理数据等
            cdr_files_map = self.get_cdr_files_map(self.folder1)
            # # 打印字典内容
            # for file_name, file_path in cdr_files_map.items():
            #     print(f"文件名: {file_name}, 文件路径: {file_path}")
            # self.read_excel(order_file=self.order_file)
            # key 为excel第一行的值
            excel_data = self.read_excel(order_file=self.order_file)
            for row_data in excel_data.values():
                try:
                    self.handleRow(cdr_files_map,row_data)
                except Exception as e:
                    self.appendRow(row_data,self.error_package)
            if self.copy_true:
                # 自动放大cdr文件
                cdr_base_white_path = os.path.join(self.folder2, '白底款')
                cdr_base_transparent_path = os.path.join(self.folder2, '透明款')
                self.magnify_cdr(cdr_base_white_path)
                self.magnify_cdr(cdr_base_transparent_path)
                self.delete_backup_cdr_files(self.folder2)
            # 完成后，发送信号到主线程
            self.finished_signal.emit("处理完毕!")
        except Exception as e:
            self.finished_signal.emit(f"处理失败: {str(e)}")

    def get_cdr_files(self,directory):
        cdr_files = []  # 用来存储符合条件的文件路径
        for root, dirs, files in os.walk(directory):  # 遍历目录及其子目录
            for file in files:
                if file.endswith('.cdr'):  # 判断文件后缀是否为 .cdr
                    cdr_files.append(os.path.join(root, file))  # 添加完整路径到列表中
        return cdr_files
    def magnify_cdr(self,cdr_base_path):

        # 顺序放大cdr
        # 遍历该目录下的所有子目录
        if os.path.exists(cdr_base_path):
            # 获取所有子目录，并过滤出数字目录
            subdirs = [item for item in os.listdir(cdr_base_path)
                       if os.path.isdir(os.path.join(cdr_base_path, item)) and item.isdigit()]

            # 将目录名转换为整数并倒序排序
            subdirs.sort(key=int, reverse=True)

            # 按照倒序排序后的目录遍历
            for subdir in subdirs:
                item_path = os.path.join(cdr_base_path, subdir)

                # 判断是否是子目录
                if os.path.isdir(item_path):
                    # 批量放大
                    cdr_files = self.get_cdr_files(item_path)
                    for cdr_file in cdr_files:
                        self.handle_cdr(cdr_file,int(subdir))
                    self.delete_backup_cdr_files(item_path)
                else:
                    print(f"文件: {subdir}")
        else:
            print("目录不存在")

    def handle_cdr(self, cdr_file , selected_size):
        try:
            # 连接 CorelDRAW 应用
            corel = win32com.client.Dispatch("CorelDRAW.Application")
            corel.Visible = False  # 不显示CorelDRAW界面，可以设置为True查看
            doc = corel.OpenDocument(cdr_file)
            # 获取页面上的所有对象
            page = doc.Pages(1)  # 获取第一页
            shapes = page.Shapes

            # 清空所有选择（通过取消选中所有图形）
            for shape in shapes:
                shape.Selected = False

            # 选择所有对象
            for shape in shapes:
                shape.Selected = True

            # 获取当前选中的对象
            selection = corel.ActiveSelection

            # 创建一个组
            group = selection.Group()
            # 获取组合的组的宽度和高度
            group_width, group_height = self.get_shape_size_in_units(corel,group)
            print(f"组合的组宽度: {group_width}, 高度: {group_height}")
            width_or_height = self.get_value_based_on_threshold(group_width,group_height)
            if width_or_height:
                self.change_width(doc,group_width, group_height, selected_size*10)
            else:
                self.change_length(doc,group_width, group_height, selected_size*10)
            # 保存为新的路径
            doc.Save()
            doc.Close()
        except Exception as e:
            # print(f"发生错误: {e}")
            print("发生错误:", e)

    # true 表示变更宽，false 表示变更长
    def get_value_based_on_threshold(self, group_width, group_height, threshold_ratio=0.05):
        # 计算两个值的差值
        difference = abs(group_width - group_height)

        # 计算差值与最大值的比例
        max_value = max(group_width, group_height)
        ratio = difference / max_value

        # 根据比例判断返回较小值还是较大值
        if ratio < threshold_ratio:
            # 比例小于阈值，返回较小值，判断是长还是宽
            if group_width < group_height:
                return True
            else:
                return False
        else:
            # 比例大于或等于阈值，返回较大值，判断是长还是宽
            if group_width > group_height:
                return True
            else:
                return False
    def change_length(self,doc,original_width, original_height, new_height):

        # 计算缩放比例（目标高度 / 原始高度）
        scale_factor = new_height / original_height

        # 计算新的宽度
        new_width = original_width * scale_factor

        # 获取解散后的所有形状（shapes）
        shapes = doc.ActivePage.Shapes
        for shape in shapes:
            shape.SetSize(shape.SizeWidth * scale_factor, shape.SizeHeight * scale_factor)

        # 输出调整后的宽度和高度（单位：mm）
        print(f"原始宽度：{original_width} mm, 原始高度：{original_height} mm")
        print(f"缩放后的宽度：{new_width} mm, 缩放后的高度：{new_height} mm")

    def change_width(self,doc,original_width, original_height, new_width):

        # 计算缩放比例（目标高度 / 原始高度）
        scale_factor = new_width / original_width

        # 计算新的宽度
        new_height = original_height * scale_factor

        # 获取解散后的所有形状（shapes）
        shapes = doc.ActivePage.Shapes
        for shape in shapes:
            shape.SetSize(shape.SizeWidth * scale_factor, shape.SizeHeight * scale_factor)

        # 输出调整后的宽度和高度（单位：mm）
        print(f"原始宽度：{original_width} mm, 原始高度：{original_height} mm")
        print(f"缩放后的宽度：{new_width} mm, 缩放后的高度：{new_height} mm")

    def get_shape_size_in_units(self,corel,shape):
        """ 获取形状的宽度和高度，并根据单位转换为合适的单位 """
        # 获取当前文档的单位
        unit = corel.ActiveDocument.Unit  # 单位可以是 mm, cm, inches, pixels 等
        width = shape.SizeWidth
        height = shape.SizeHeight

        # 英寸（inches）转换为厘米（cm）
        if unit == 1:  # 1: inches
            width *= 25.4  # 将英寸转换为厘米
            height *= 25.4
        elif unit == 2:  # 2: mm
            # width /= 10  # 将毫米转换为厘米
            # height /= 10
            pass
        elif unit == 3:  # 3: cm
            width *= 10  # 厘米转毫米
            height *= 10
        elif unit == 7:  # 7: pixels
            # 无需转换，单位已经是像素
            pass
        else:
            print(f"不支持的单位: {unit}")

        return width, height

    def delete_backup_cdr_files(self,folder_path):
        # 遍历文件夹中的所有文件
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                # 检查文件名是否以 "Backup_of_" 开头且以 ".cdr" 结尾
                if file.startswith("Backup_of_") and file.endswith(".cdr"):
                    file_path = os.path.join(root, file)
                    try:
                        # 删除文件
                        os.remove(file_path)
                        print(f"Deleted: {file_path}")
                    except Exception as e:
                        print(f"Failed to delete {file_path}: {e}")

    def appendRow(self, row_data,file_path):
        if os.path.exists(file_path):
            # 文件已存在，打开已有文件
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            # 文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            sheet = wb.active
            # 写入标题行（假设 data 的第一个字典是列名）
            headers = list(row_data.keys())
            sheet.append(headers)
        values = list(row_data.values())
        sheet.append(values)
        # 保存文件
        wb.save(file_path)

    def is_empty_string(self,value):
        # 检查是否是字符串类型且为空字符串
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    # 处理每一行
    def handleRow(self,cdr_files_map,row_data):
        order_num = row_data.get('订单编号')
        if not order_num:
            order_num = row_data.get('订单号')
        if not order_num:
            self.appendRow(row_data, self.error_package)
            return
        specification_name_str  = row_data.get('规格名称')
        style,longest_side = self.parse_specification_name_str(specification_name_str)
        cdr_file_path = cdr_files_map.get(style)
        if not cdr_file_path:
            # 款号不存在
            self.appendRow(row_data,self.lack_package)
            return
        if not self.is_valid_longest_side(longest_side):
            # 最长边解析异常
            self.appendRow(row_data, self.error_package)
            return
        # 买家留言
        if not self.is_empty_string(row_data.get('备注')) or not self.is_empty_string(row_data.get('买家留言')):
            self.appendRow(row_data, self.remain_package)
            return

        self.copy_cdr(row_data,style,longest_side,cdr_file_path)

    def copy_cdr(self,row_data,style, longest_side,cdr_file_path):
        order_num = row_data.get('订单编号')
        if not order_num:
            order_num = row_data.get('订单号')
        cdr_base_path = os.path.join(self.folder2,'白底款')
        if 'T' in style or 't' in style:
            # 透明款
            cdr_base_path =  os.path.join(self.folder2,'透明款')
        # 拷贝文件，并把文件名改成订单编号
        dst_dir = os.path.join(cdr_base_path, str(longest_side))
        # 处理多件
        good_nums = int(row_data.get('商品数量'))
        if not good_nums:
            good_nums = int(row_data.get('数量'))
        if good_nums > 1:
            dst_dir = os.path.join(dst_dir, str(good_nums))
        if good_nums == 1:
            self.copy_file_with_new_name(cdr_file_path, dst_dir, order_num)
        else:
            self.copy_file_with_new_name_nums(cdr_file_path, dst_dir, good_nums,order_num)
        cdr_excel_path = os.path.join(self.folder2, '统计数据.xlsx')
        self.appendCdrRow(row_data,style,longest_side,cdr_excel_path)

    def appendCdrRow(self, row_data,style, longest_side, file_path):
        num = row_data.get('数量')
        if not num:
            num = row_data.get('商品数量')
        order_num =  row_data.get('订单编号')
        if not order_num:
            order_num = row_data.get('订单号')
        new_row_data  = {
            '订单编号': order_num,
            '店铺名称': row_data.get('店铺名称'),
            '规格名称': row_data.get('规格名称'),
            '规格':style,
            '最长边':longest_side,
            '数量':num,
            '总价': row_data.get('总价'),
            '实收': row_data.get('实收')

        }
        if os.path.exists(file_path):
            # 文件已存在，打开已有文件
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
        else:
            # 判断目标目录是否存在，如果不存在则创建
            if not os.path.exists(os.path.dirname(file_path)):
                os.makedirs(os.path.dirname(file_path))
            # 文件不存在，创建一个新的工作簿
            wb = openpyxl.Workbook()
            sheet = wb.active
            # 写入标题行（假设 data 的第一个字典是列名）
            headers = list(new_row_data.keys())
            sheet.append(headers)
        values = list(new_row_data.values())
        sheet.append(values)
        # 保存文件
        wb.save(file_path)

    def copy_file_with_new_name_nums(self,src_file, dst_dir, good_nums , new_name=None):
        # 获取源文件的文件名和扩展名
        src_filename, file_extension = os.path.splitext(os.path.basename(src_file))

        # 如果没有提供新的文件名，则使用原文件名
        if new_name is None:
            new_name = src_filename  # 保持原文件名的基础部分

        for idx, num in enumerate(range(1, good_nums + 1), start=1):
            if idx == 1:  # 第一个元素命名为 xxx
                # 构建目标文件名，保持原有扩展名
                new_file_name = new_name + file_extension

                # 构建目标文件路径
                dst_file = os.path.join(dst_dir, new_file_name)

                # 判断目标目录是否存在，如果不存在则创建
                if not os.path.exists(dst_dir):
                    os.makedirs(dst_dir)

                # 拷贝文件
                shutil.copy(src_file, dst_file)
                print(f"文件 '{src_file}' 已成功拷贝到 '{dst_file}'.")
            else:  # 后续元素命名为 xxx_2, xxx_3, ...
                # 构建目标文件名，保持原有扩展名
                new_file_name = new_name+'_'+str(idx) + file_extension

                # 构建目标文件路径
                dst_file = os.path.join(dst_dir, new_file_name)

                # 判断目标目录是否存在，如果不存在则创建
                if not os.path.exists(dst_dir):
                    os.makedirs(dst_dir)

                # 拷贝文件
                shutil.copy(src_file, dst_file)
                print(f"文件 '{src_file}' 已成功拷贝到 '{dst_file}'.")

    def copy_file_with_new_name(self,src_file, dst_dir, new_name=None):
        # 获取源文件的文件名和扩展名
        src_filename, file_extension = os.path.splitext(os.path.basename(src_file))

        # 如果没有提供新的文件名，则使用原文件名
        if new_name is None:
            new_name = src_filename  # 保持原文件名的基础部分

        # 构建目标文件名，保持原有扩展名
        new_file_name = new_name + file_extension

        # 构建目标文件路径
        dst_file = os.path.join(dst_dir, new_file_name)

        # 判断目标目录是否存在，如果不存在则创建
        if not os.path.exists(dst_dir):
            os.makedirs(dst_dir)

        # 拷贝文件
        shutil.copy(src_file, dst_file)
        print(f"文件 '{src_file}' 已成功拷贝到 '{dst_file}'.")

    def is_valid_longest_side(self,longest_side):
        # 定义允许的数值集合
        valid_values = {15, 20, 30, 40, 50, 60, 70, 80, 90, 100}

        # 判断 longest_side 是否是数字
        if isinstance(longest_side, (int, float)):  # 如果是整数或浮点数
            return longest_side in valid_values
        elif isinstance(longest_side, str) and longest_side.isdigit():  # 如果是字符串且为数字
            return int(longest_side) in valid_values
        else:
            return False

    def parse_specification_name_str(self,specification_name_str):
        # 提取款号 (假设款号是字母+数字的组合，可以匹配如 CD115-A)
        match_model = re.search(r'([A-Za-z]+\d+-[A-Za-z])', specification_name_str)
        # 提取所有尺寸，可能的格式为：40cm高x35cm宽 或 60x60cm
        size_matches = re.findall(r'(\d+)(?=cm|x|CM|厘米|\*|X|公分)', specification_name_str)
        if match_model and size_matches:
            model_number = match_model.group(1)
            # 转换尺寸为整数并找到最大的尺寸作为最长边
            sizes = [int(size) for size in size_matches]
            max_size = max(sizes)
        return model_number,max_size

    # 读取订单excel
    def read_excel(self,order_file):
        workbook = openpyxl.load_workbook(order_file)
        # 选择活动工作表
        sheet = workbook.active
        # 读取第一行作为键（keys）
        header = [cell.value for cell in sheet[1]]  # 第一行作为键
        # 读取数据（从第二行开始）
        data = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {header[i]: row[i] for i in range(len(header))}
            data[row_data[header[0]]] = row_data  # 使用第一列的值作为外层字典的 key
        return data

    def is_cdr_file(self,file_path):
        """检查文件是否为 .cdr 格式"""
        return file_path.lower().endswith('.cdr')

    def get_cdr_files_map(self,directory):
        """遍历目录，返回文件名为键，文件路径为值的字典"""
        cdr_files_map = {}

        # 遍历指定目录下的所有文件和文件夹
        for root, dirs, files in os.walk(directory):
            for file in files:
                file_path = os.path.join(root, file)

                # 如果是 .cdr 文件，则将文件名和路径存入字典
                if self.is_cdr_file(file_path):
                    file_name =os.path.splitext(os.path.basename(file))[0]  # 获取文件名（不含路径）
                    cdr_files_map[file_name] = file_path

        return cdr_files_map




class FolderSelector(QWidget):
    def __init__(self):
        super().__init__()

        self.init_ui()

    def init_ui(self):
        # 创建控件
        self.label1 = QLabel('请选择图库文件夹:')
        self.label2 = QLabel('请选择输出文件夹（建议文件夹内容为空）:')

        self.folder1_path = QLabel('未选择文件夹')
        self.folder2_path = QLabel('未选择文件夹')
        # self.folder1_path = QLabel('//xinguo/贴纸.生产资料/美工做的新款，待检查')
        # self.folder2_path = QLabel('E:/导出')

        self.select_folder1_btn = QPushButton('图库文件夹')
        self.select_folder2_btn = QPushButton('输出文件夹')
        self.ok_btn = QPushButton('确定')
        self.cancel_btn = QPushButton('取消')

        # True/False 单选框：是否拷贝文件
        self.copy_group = QGroupBox("是否自动放大文件", self)
        self.copy_true = QRadioButton("是 (自动放大cdr文件)", self)
        self.copy_false = QRadioButton("否 (不放大cdr文件)", self)
        self.copy_false.setChecked(True)  # 默认选择“否”

        # 设置单选框的布局
        copy_layout = QHBoxLayout()
        copy_layout.addWidget(self.copy_true)
        copy_layout.addWidget(self.copy_false)
        self.copy_group.setLayout(copy_layout)

        # 按钮点击事件
        self.select_folder1_btn.clicked.connect(self.select_folder1)
        self.select_folder2_btn.clicked.connect(self.select_folder2)
        self.ok_btn.clicked.connect(self.ok_clicked)
        self.cancel_btn.clicked.connect(self.cancel_clicked)

        # 布局设置
        layout = QVBoxLayout()

        # 第一个文件夹选择行
        folder1_layout = QHBoxLayout()
        folder1_layout.addWidget(self.label1)
        folder1_layout.addWidget(self.folder1_path)
        folder1_layout.addWidget(self.select_folder1_btn)
        layout.addLayout(folder1_layout)

        # 第二个文件夹选择行
        folder2_layout = QHBoxLayout()
        folder2_layout.addWidget(self.label2)
        folder2_layout.addWidget(self.folder2_path)
        folder2_layout.addWidget(self.select_folder2_btn)
        layout.addLayout(folder2_layout)

        layout.addWidget(self.copy_group)  # 添加是否拷贝文件的单选框

        # 创建按钮：选择文件
        self.open_button = QPushButton('订单列表，请选择 .xlsx 文件')
        self.open_button.clicked.connect(self.open_file_dialog)
        # 创建标签，用于显示选择的文件路径
        self.order_file = QLabel('未选择文件', self)
        # self.order_file = QLabel('E:/627个.xlsx', self)
        # 将按钮和标签加入布局
        layout.addWidget(self.open_button)
        layout.addWidget(self.order_file)

        # 确定和取消按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.ok_btn)
        buttons_layout.addWidget(self.cancel_btn)
        layout.addLayout(buttons_layout)

        self.setLayout(layout)

        # 窗口设置
        self.setWindowTitle('订单分拣-zc')
        self.setGeometry(200, 200, 600, 150)

    def open_file_dialog(self):
        # 打开文件选择框，限制只选择 .xlsx 文件
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "选择 .xlsx 文件", "", "Excel 文件 (*.xlsx);;所有文件 (*)",
                                                   options=options)

        if file_name:
            # 显示选择的文件路径
            self.order_file.setText(f"{file_name}")
        else:
            self.order_file.setText("未选择文件")

    def select_folder1(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择文件夹 1')
        if folder_path:
            self.folder1_path.setText(folder_path)

    def select_folder2(self):
        folder_path = QFileDialog.getExistingDirectory(self, '选择文件夹 2')
        if folder_path:
            self.folder2_path.setText(folder_path)

    def ok_clicked(self):
        # 获取选择的文件夹路径
        folder1 = self.folder1_path.text()
        folder2 = self.folder2_path.text()
        order_file = self.order_file.text()
        if not folder1 or not folder2 or not os.path.exists(folder1) or not os.path.exists(folder2):
            QMessageBox.warning(self, '警告', '请确保图库文件夹和导出文件夹都已经选择且正确！')
            return
        if not order_file or not os.path.isfile(order_file):
            QMessageBox.warning(self, '警告', '订单excel选择有误！')
            return
        # 启动后台线程执行任务
        self.thread = LoadThread(folder1, folder2,order_file,self.copy_true.isChecked())
        self.thread.finished_signal.connect(self.on_load_finished)
        self.thread.start()

        # 禁用按钮以防止重复点击
        self.ok_btn.setDisabled(True)
        self.cancel_btn.setDisabled(True)

    def cancel_clicked(self):
        # 创建确认对话框
        reply = QMessageBox.question(self, '确认取消', '你确定要取消操作吗?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        print("操作已取消")
        self.close()

    def on_load_finished(self, message):
        # 加载完成后弹出提示框
        QMessageBox.information(self, '提示', message)

        # 启用按钮
        self.ok_btn.setEnabled(True)
        self.cancel_btn.setEnabled(True)

        # 如果操作完成，关闭窗口
        self.close()

def main():
    app = QApplication(sys.argv)
    window = FolderSelector()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
